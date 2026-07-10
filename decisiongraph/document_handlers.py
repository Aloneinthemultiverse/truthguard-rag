import os
import csv
import json


def read_txt(path: str) -> str:
    with open(path, "r", errors="ignore") as f:
        text = f.read()
    print(f"  Total chars: {len(text)}")
    return text


def read_md(path: str) -> str:
    return read_txt(path)


def read_docx(path: str) -> str:
    try:
        import docx
        doc = docx.Document(path)
        text = "\n".join([para.text for para in doc.paragraphs if para.text.strip()])
        print(f"  Total chars: {len(text)}")
        return text
    except ImportError:
        raise ImportError("python-docx not installed. Run: pip install python-docx")


def read_csv(path: str) -> str:
    rows = []
    with open(path, "r", errors="ignore") as f:
        reader = csv.DictReader(f)
        for row in reader:
            # convert each row to natural language sentence
            row_text = " | ".join([f"{k}: {v}" for k, v in row.items() if v])
            rows.append(row_text)
    text = "\n".join(rows)
    print(f"  Total chars: {len(text)} ({len(rows)} rows)")
    return text


def read_excel(path: str) -> str:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True)
        all_text = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            all_text.append(f"Sheet: {sheet_name}")
            headers = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(c) if c else f"col{j}" for j, c in enumerate(row)]
                    continue
                if any(c is not None for c in row):
                    row_text = " | ".join([
                        f"{headers[j]}: {c}"
                        for j, c in enumerate(row)
                        if c is not None and j < len(headers)
                    ])
                    all_text.append(row_text)
        text = "\n".join(all_text)
        print(f"  Total chars: {len(text)}")
        return text
    except ImportError:
        raise ImportError("openpyxl not installed. Run: pip install openpyxl")


def read_email(path: str) -> str:
    import email
    with open(path, "r", errors="ignore") as f:
        msg = email.message_from_file(f)
    parts = []
    parts.append(f"From: {msg.get('From', '')}")
    parts.append(f"To: {msg.get('To', '')}")
    parts.append(f"Subject: {msg.get('Subject', '')}")
    parts.append(f"Date: {msg.get('Date', '')}")
    if msg.is_multipart():
        for part in msg.walk():
            if part.get_content_type() == "text/plain":
                parts.append(part.get_payload(decode=True).decode("utf-8", errors="ignore"))
    else:
        parts.append(msg.get_payload(decode=True).decode("utf-8", errors="ignore"))
    text = "\n".join(parts)
    print(f"  Total chars: {len(text)}")
    return text


def read_document(path: str) -> str:
    ext = os.path.splitext(path)[1].lower()
    handlers = {
        ".pdf":  _read_pdf,
        ".txt":  read_txt,
        ".md":   read_md,
        ".docx": read_docx,
        ".csv":  read_csv,
        ".xlsx": read_excel,
        ".xls":  read_excel,
        ".eml":  read_email,
    }
    handler = handlers.get(ext)
    if not handler:
        raise ValueError(f"Unsupported file type: {ext}")
    return handler(path)


def _read_pdf(path: str) -> str:
    from .ingest import read_pdf
    return read_pdf(path)
